import platform
import re
import subprocess
from datetime import datetime
from openpyxl import Workbook
import time
from appium.webdriver.common.appiumby import AppiumBy

from framework_constants import Ios_franklin_Constants
from appium.webdriver.common.appiumby import AppiumBy
from selenium.common.exceptions import NoSuchElementException
from openpyxl.styles import Side, Alignment
from openpyxl.styles import PatternFill, Font, Border
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.formatting.rule import Rule
from openpyxl import Workbook
from openpyxl.chart import PieChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart.label import DataLabelList

def get_current_time():
    """
    This function returns current time
    :return:
    """
    current_time = datetime.now()
    return current_time

def execution_time(tc_start_time):
    """
        This function returns overall execution time
        :return:
    """
    tc_end_time = get_current_time()
    tc_time_interval = tc_end_time - tc_start_time
    tc_time_interval_str = str(tc_time_interval).split(".")[0]
    print("****************Execution completed***********************")

# def get_device():
#     global device_id
#     device_id = get_iOS_devices()
#     print("printing device id ::::",device_id)
# #------------------------------------------------------------
# def get_iOS_devices():
#
#     ios_emulator_output = subprocess.Popen('xcrun simctl list devices | grep "(Booted)"', stdout=subprocess.PIPE,
#                                            shell=True)
#     output2 = ios_emulator_output.stdout.read().strip().decode()
#     ios_emulator_output.stdout.close()
#     ios_emulator = []
#     if output2:
#         # ios_emulator_output.terminate()
#         out = output2.split("\n")
#         for i in out:
#             dev_output1 = re.search(r"\b([0-9a-fA-F]{8}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{4}-[0-9a-fA-F]{12})\b", i)
#             if dev_output1:
#                 dev_list1 = dev_output1.group(0)
#                 ios_emulator.append(dev_list1.strip())
#
#     return ios_emulator
#----------------------------------------------------------

# def get_iOS_devices():
#     ios_device_output = subprocess.Popen('idevice_id -l', stdout=subprocess.PIPE,
#                                                shell=True)
#     output = ios_device_output.stdout.read().strip().decode()
#     device_out = ios_device_output.stderr.read()
#     print(device_out)
#     ios_device_output.stdout.close()
#     ios_device = []
#     if output:
#         #ios_emulator_output.terminate()
#         out = output.split("\n")
#         for i in out:
#             dev_output = re.search(r"[a-z,A-z,\d,\s,-]+", i)
#             if dev_output:
#
#                 dev_list = dev_output.group(0)
#                 ios_device.append(dev_list.strip())
#     print(ios_device)
#     return ios_device

def get_iOS_devices():
    result = subprocess.run("ideviceinfo | grep -i UniqueDeviceID", shell=True, capture_output=True, text=True)
    product_version = result.stdout.rstrip()
    version = product_version.find(":")
    print(product_version[version + 2:])
    ios_device = product_version[version + 2:]
    print(ios_device,"sindhi")
    return ios_device


# -------------------------- Installing .app for iOS Simulator -------------------

app_path = "/Users/nadabala/Downloads/Convatec.app"

def checking_app_installation(serial):
    try:
        cmd = f'xcrun simctl listapps {serial}'
        process = subprocess.Popen(cmd, stdout=subprocess.PIPE, shell=True)
        apps = process.stdout.read().strip().decode()
        if "FranklinDev" in apps:
            print("Franklin application is already installed")
        else:
            print("")
            print("Franklin application is not installed installing ..........")
            install_cmd = f'xcrun simctl install {serial} {app_path}'
            print("")
            process_install = subprocess.Popen(install_cmd, stdout=subprocess.PIPE, shell=True)
            cmd = f'xcrun simctl listapps {serial}'
            process = subprocess.Popen(cmd, stdout=subprocess.PIPE, shell=True)
            print('Wait untill franklin application installation ....')
            time.sleep(4)
            apps = process.stdout.read().strip().decode()
            if "com.convatec.franklindev" in apps:
                print("Franklin application is installed")
    except Exception as error:
        print(error)



#-------------------------- Checking appium launched or not ----------------------------


# Operating system

WINDOWS = "Windows"
iOS = "Darwin"
def operating_system():
    """
    This function returns the home directory of the machine
    :return:
    """
    if platform.system() == iOS:
        return iOS
    elif platform.system() == WINDOWS:
        return WINDOWS
    else:
        return platform.system()

OPERATING_SYSTEM = operating_system()





def Checking_appium():
    cmd_windows = 'netstat -an | findstr /C:"4723"'
    cmd_linux = "lsof -i :4723"
    if OPERATING_SYSTEM == WINDOWS:
        process_windows = subprocess.getoutput(cmd_windows)
        if process_windows:
            return True
        else:
            return False
    elif OPERATING_SYSTEM == iOS:
        proocess_linux = subprocess.getoutput(cmd_linux)
        if proocess_linux:
            print("")
            print("Appium is running in port number : 4723")
            print("")
            return True
        else:
            print("Please check Appium is running in port number : 4723")
            return False
    else:
        return False
#00008130-001858402EF8001C


def desired_caps_IOS(serial):
    from appium import webdriver
    from appium.options.common import AppiumOptions
    print("Serail is ::::",serial)
    desired_caps = {}
    desired_caps['platformName'] = "iOS"
    desired_caps['deviceName'] = "iPhone 15 Plus"
    desired_caps['udid'] = serial
    desired_caps['bundleId'] = "com.convatec.franklindev"
    desired_caps['platformVersion'] = "17.2"
    desired_caps[''] = "XCUITest"
    # desired_caps['noReset'] = True
    # desired_caps['fullReset'] = False
    desired_caps['showIOSLog'] = True
    desired_caps['newCommandTimeout'] = 340
    option = AppiumOptions().load_capabilities(desired_caps)
    driver = webdriver.Remote(command_executor=f"http://localhost:4723/wd/hub", options=option)
    # print(driver)
    # sender_driver = webdriver.Remote('http://localhost:4723/wd/hub',desired_caps)
    driver.implicitly_wait(6)
    print("IOS driver initialized")
    return driver



# def create_workbook():
#     """
#     This function create openpyxl workbook, test summary sheet and test report sheet objects
#     :param sheet1: sheet name of test report
#     :return: returns openpyxl workbook, test summary sheet and test report sheet objects
#     """
#     report_wokbook = Workbook()
#     summary_sheet = report_wokbook.active
#     summary_sheet.title = 'Test Summary'
#     new_summary_sheet = report_wokbook.create_sheet(title='Overall Summary')
#     top_column = ("AZURE ID", "Testcases Function", "Status", "Reason")
#     summary_sheet.append(top_column)
#     return report_wokbook, summary_sheet, new_summary_sheet
#
# def create_summary(summary_sheet, execution_data):
#     """
#     :param ws2: summary sheet object
#     :param data: testcase execution data
#     :param execution_time: execution start and end time
#     :param test_type: test type to create pie chart
#     :return: None
#     """
#     test_type_count = 0
#     row_space = 12
#     pie_chart_row_space = 16
#     print(execution_data)
#     execution_status_data = execution_data
#     # execution_time = execution_data[1]
#     execution_status = [
#         ['Description', 'Count'],
#         [f'Total Test Cases', execution_status_data[0]],
#         ['Passed', execution_status_data[1]],
#         ['Failed', execution_status_data[2]],
#         ['Not Executed', execution_status_data[3]]
#     ]
#
#     # execution_timings = [
#     #     ['', ''],
#     #     [f'Execution Details for', ''],
#     #     ['Execution Start Time', execution_time[0]],
#     #     ['Execution End Time', execution_time[1]],
#     #     ['Total Duration', execution_time[1] - execution_time[0]]
#     # ]
#
#     gap_cells = [
#         ['', ''],
#         ['', ''],
#     ]
#
#     summary_sheet.column_dimensions['A'].width = 25
#     summary_sheet.column_dimensions['B'].width = 20
#
#     for row in execution_status:
#         summary_sheet.append(row)
#
#     # for row in execution_timings:
#     #     summary_sheet.append(row)
#
#     for row in gap_cells:
#         summary_sheet.append(row)
#
#     pie = PieChart()
#     labels = Reference(summary_sheet, min_col=1, min_row=3 + test_type_count * row_space,
#                        max_row=5 + test_type_count * row_space)
#     data = Reference(summary_sheet, min_col=2, min_row=2 + test_type_count * row_space,
#                      max_row=5 + test_type_count * row_space)
#     pie.add_data(data, titles_from_data=True)
#     pie.set_categories(labels)
#     pie.title = f"Test Summary"
#
#     summary_sheet.cell(row=1 + test_type_count * row_space, column=1).font = Font(b=True)
#     summary_sheet.cell(row=1 + test_type_count * row_space, column=2).font = Font(b=True)
#     summary_sheet.cell(row=7 + test_type_count * row_space, column=1).font = Font(b=True)
#     # summary_sheet.cell(row=12, column=1).font = Font(b=True)
#
#     series = pie.series[0]
#     pt = DataPoint(idx=0)
#     pt.graphicalProperties.solidFill = "238823"  # G
#     series.dPt.append(pt)
#
#     pt = DataPoint(idx=1)
#     pt.graphicalProperties.solidFill = "D2222D"  # R
#     series.dPt.append(pt)
#
#     pt = DataPoint(idx=2)
#     pt.graphicalProperties.solidFill = "0099E5"  # Y
#     series.dPt.append(pt)
#
#     # Showing data labels as percentage
#     starting_cell = "E" + str(2 + test_type_count * pie_chart_row_space)
#     pie.dataLabels = DataLabelList()
#     pie.dataLabels.showPercent = True
#     summary_sheet.add_chart(pie, starting_cell)
#     print(execution_status_data)
#
# def update_status(az_id,func_data, status,status_reason, sheet):
#     """
#     This function update the result in excel report
#     :param funcname: name of the funcion
#     :param func_data: dictionary of (testcase_data) pairs
#     :param status_reason: reason for execution started/not started
#     :param sheet: excel report test report
#     :return:
#     """
#     data = (
#         az_id, func_data,status, status_reason)
#     sheet.append(data)
#
#
# def get_execution_status(sheet):
#     """
#     This function returns testexecution status for a test_type
#     :param sheet: sheet object
#     :param device_serial: device serial
#     :param test_type: test_type
#     :return: test execution data
#     """
#     runExcelRules(sheet)
#     status = get_status_list(sheet)
#     total_cases = len(status)
#     passed_cases = status.count('Pass')
#     failed_cases = status.count('Fail')
#     not_executed_cases = total_cases - (passed_cases + failed_cases)
#     print("*******get_exec_status_excel_report**********")
#     execution_status_data = [total_cases, passed_cases, failed_cases, not_executed_cases]
#     return execution_status_data
#
#
#
# def get_status_list(sheet):
#     """
#     This function returns list of test exceution status
#     :param sheet: test report sheet
#     :return: read the pass/fail criteria cells and return the test execution status
#     """
#     status_list = []
#     for value in sheet.iter_rows(min_row=1,
#                                  max_row=10000,
#                                  min_col=3, max_col=3, values_only=True):
#         if value[0] == "Pass":
#             status_list.append(value[0])
#         elif value[0] == "Fail":
#             status_list.append(value[0])
#         elif value[0] == "Not Executed":
#             status_list.append(value[0])
#
#     return status_list
#
# def runExcelRules(ws=None):
#     """
#     This function add colors to pass/fail text and changes alignment for testreport cells
#     :param ws: sheet name
#     :return: None
#     """
#     # Highlight cells that contain particular text by using a special formula
#     black_text = Font(color="000000")  # "9C0006"
#     red_fill = PatternFill(bgColor="C13307")  # "FFC7CE"
#     green_fill = PatternFill(bgColor="1AAB07")  # "FFC7CE"
#
#     dxf = DifferentialStyle(font=black_text, fill=red_fill)
#     rule = Rule(type="containsText", operator="containsText", dxf=dxf)  # text="Microsoft",
#     rule.formula = ['NOT(ISERROR(SEARCH("Fail",E1)))']
#     ws.conditional_formatting.add('E1:E200', rule)
#
#     dxf = DifferentialStyle(font=black_text, fill=green_fill)
#     rule = Rule(type="containsText", operator="containsText", dxf=dxf)  # text="Microsoft",
#     rule.formula = ['NOT(ISERROR(SEARCH("Pass",E1)))']
#     ws.conditional_formatting.add('E1:E200', rule)
#
#     # Wrap Text
#     for column in ws.iter_cols(min_col=4):
#         for cell in column:
#             cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')
#
#     for column in ws.iter_cols(min_col=5):
#         for cell in column:
#             cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')
#
#     for column in ws.iter_cols(min_col=6):
#         for cell in column:
#             cell.alignment = Alignment(wrap_text=True, vertical='bottom', horizontal='left')
#
# def get_execution_status(sheet):
#     """
#     This function returns testexecution status for a test_type
#     :param sheet: sheet object
#     :param device_serial: device serial
#     :param test_type: test_type
#     :return: test execution data
#     """
#     runExcelRules(sheet)
#     status = get_status_list(sheet)
#     total_cases = len(status)
#     passed_cases = status.count('Pass')
#     failed_cases = status.count('Fail')
#     not_executed_cases = total_cases - (passed_cases + failed_cases)
#     print("*******get_exec_status_excel_report**********")
#     execution_status_data = [total_cases, passed_cases, failed_cases, not_executed_cases]
#     return execution_status_data

def create_workbook():
    """
    This function create openpyxl workbook, test summary sheet and test report sheet objects
    :param sheet1: sheet name of test report
    :return: returns openpyxl workbook, test summary sheet and test report sheet objects
    """
    report_wokbook = Workbook()
    summary_sheet = report_wokbook.active
    summary_sheet.title = 'Test Summary'
    return report_wokbook,summary_sheet

def create_report_sheet(workbook=""):
    """
    This function create openpyxl workbook, test summary sheet and test report sheet objects
    :param sheet1: sheet name of test report
    :return: returns openpyxl workbook, test summary sheet and test report sheet objects
    """
    ws = workbook.create_sheet(f"Franklin TestCases Execution Report")
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 50
    # ws.column_dimensions['E'].width = 20
    # ws.column_dimensions['F'].width = 50

    # Document Title
    ws.merge_cells('A1:E2')
    top_cell = ws['A1']
    top_merged = ws['A1:E2']
    top_cell2 = ws['A3']
    top_cell.value = "Test Report"

    thin = Side(border_style="thin", color="000000")

    for each in list(top_merged):
        for each2 in each:
            each2.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    top_cell.fill = PatternFill("solid", fgColor="FF9900")
    top_cell.font = Font(name='Calibri', b=True, color="000000", size="16")
    top_cell.alignment = Alignment(horizontal="center", vertical="center")
    top_cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
    top_column = ("AZURE ID","Testcases Function","Status","Reason")
    ws.append(top_column)
    ws.cell(row=4, column=1).font = Font(bold=True)
    ws.cell(row=4, column=1).border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws.cell(row=4, column=2).font = Font(bold=True)
    ws.cell(row=4, column=2).border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws.cell(row=4, column=3).font = Font(bold=True)
    ws.cell(row=4, column=3).border = Border(top=thin, left=thin, right=thin, bottom=thin)
    ws.cell(row=4, column=4).font = Font(bold=True)
    ws.cell(row=4, column=4).border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # ws.cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    return ws

def create_summary(summary_sheet, execution_data):
    """
    :param ws2: summary sheet object
    :param data: testcase execution data
    :param execution_time: execution start and end time
    :param test_type: test type to create pie chart
    :return: None
    """
    thin = Side(border_style="thin", color="000000")
    test_type_count = 0
    row_space = 12
    pie_chart_row_space = 16
    print(execution_data)
    execution_status_data = execution_data
    # execution_time = execution_data[1]
    execution_status = [
        ['Description', 'Count'],
        [f'Total Test Cases', execution_status_data[0]],
        ['Passed', execution_status_data[1]],
        ['Failed', execution_status_data[2]],
        ['Not Executed', execution_status_data[3]]
    ]

    # execution_timings = [
    #     ['', ''],
    #     [f'Execution Details for', ''],
    #     ['Execution Start Time', execution_time[0]],
    #     ['Execution End Time', execution_time[1]],
    #     ['Total Duration', execution_time[1] - execution_time[0]]
    # ]

    gap_cells = [
        ['', ''],
        ['', ''],
    ]

    summary_sheet.column_dimensions['A'].width = 25
    summary_sheet.column_dimensions['B'].width = 20

    for row in execution_status:
        summary_sheet.append(row)

    for row in summary_sheet.iter_rows():
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

    # for row in execution_timings:
    #     summary_sheet.append(row)

    for row in gap_cells:
        summary_sheet.append(row)

    pie = PieChart()
    labels = Reference(summary_sheet, min_col=1, min_row=3 + test_type_count * row_space,
                       max_row=5 + test_type_count * row_space)
    data = Reference(summary_sheet, min_col=2, min_row=2 + test_type_count * row_space,
                     max_row=5 + test_type_count * row_space)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = f"Test Summary"

    summary_sheet.cell(row=1 + test_type_count * row_space, column=1).font = Font(b=True)
    summary_sheet.cell(row=1 + test_type_count * row_space, column=2).font = Font(b=True)
    summary_sheet.cell(row=7 + test_type_count * row_space, column=1).font = Font(b=True)
    # summary_sheet.cell(row=12, column=1).font = Font(b=True)

    series = pie.series[0]
    pt = DataPoint(idx=0)
    pt.graphicalProperties.solidFill = "238823"  # G
    series.dPt.append(pt)

    pt = DataPoint(idx=1)
    pt.graphicalProperties.solidFill = "D2222D"  # R
    series.dPt.append(pt)

    pt = DataPoint(idx=2)
    pt.graphicalProperties.solidFill = "0099E5"  # Y
    series.dPt.append(pt)

    # Showing data labels as percentage
    starting_cell = "E" + str(2 + test_type_count * pie_chart_row_space)
    pie.dataLabels = DataLabelList()
    pie.dataLabels.showPercent = True
    summary_sheet.add_chart(pie, starting_cell)
    print(execution_status_data)

def update_status(az_id,func_data, status,status_reason, sheet):
    """
    This function update the result in excel report
    :param funcname: name of the funcion
    :param func_data: dictionary of (testcase_data) pairs
    :param status_reason: reason for execution started/not started
    :param sheet: excel report test report
    :return:
    """
    thin = Side(border_style="thin", color="000000")
    data = (
        az_id,func_data,status, status_reason)
    sheet.append(data)
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)

def get_execution_status(sheet):
    """
    This function returns testexecution status for a test_type
    :param sheet: sheet object
    :param device_serial: device serial
    :param test_type: test_type
    :return: test execution data
    """
    runExcelRules(sheet)
    status = get_status_list(sheet)
    total_cases = len(status)
    passed_cases = status.count('Pass')
    failed_cases = status.count('Fail')
    not_executed_cases = total_cases - (passed_cases + failed_cases)
    print("*******get_exec_status_excel_report**********")
    execution_status_data = [total_cases, passed_cases, failed_cases, not_executed_cases]
    return execution_status_data


# def generate_email():
#     mailFilePath = "C:\\Users\\sabbineni\\Desktop\\Convatec\\maildata"
#     mail_data = mailcsv_to_list_of_dicts(mailFilePath)
#     print(mail_data)
#     if mail_data:
#         idls = []
#         print("line 1098")
#         num = idls[-1]
#         pattern = r"ammu(\d+)@yopmail.com"
#         match = re.search(pattern, num)
#         if match:
#             k = match.group(1)
#             d = int(k) + 1
#             mailid = "ammu" + str(d) +"@yopmail.com"
#             print(mailid)
#             return mailid
#     else:
#         print("if block")
#         mailid="ammu" + "1" +"@yopmail.com"
#         print(mailid)
#         return mailid

def get_status_list(sheet):
    """
    This function returns list of test exceution status
    :param sheet: test report sheet
    :return: read the pass/fail criteria cells and return the test execution status
    """
    status_list = []
    for value in sheet.iter_rows(min_row=1,
                                 max_row=10000,
                                 min_col=3, max_col=3, values_only=True):
        if value[0] == "Pass":
            status_list.append(value[0])
        elif value[0] == "Fail":
            status_list.append(value[0])
        elif value[0] == "Not Executed":
            status_list.append(value[0])

    return status_list

def runExcelRules(ws=None):
    """
    This function add colors to pass/fail text and changes alignment for testreport cells
    :param ws: sheet name
    :return: None
    """
    # Highlight cells that contain particular text by using a special formula
    black_text = Font(color="000000")  # "9C0006"
    red_fill = PatternFill(bgColor="C13307")  # "FFC7CE"
    green_fill = PatternFill(bgColor="1AAB07")  # "FFC7CE"

    dxf = DifferentialStyle(font=black_text, fill=red_fill)
    rule = Rule(type="containsText", operator="containsText", dxf=dxf)  # text="Microsoft",
    rule.formula = ['NOT(ISERROR(SEARCH("Fail",E1)))']
    ws.conditional_formatting.add('E1:E200', rule)

    dxf = DifferentialStyle(font=black_text, fill=green_fill)
    rule = Rule(type="containsText", operator="containsText", dxf=dxf)  # text="Microsoft",
    rule.formula = ['NOT(ISERROR(SEARCH("Pass",E1)))']
    ws.conditional_formatting.add('E1:E200', rule)

    # Wrap Text
    for column in ws.iter_cols(min_col=4):
        for cell in column:
            cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='left')

    for column in ws.iter_cols(min_col=5):
        for cell in column:
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')

    for column in ws.iter_cols(min_col=6):
        for cell in column:
            cell.alignment = Alignment(wrap_text=True, vertical='bottom', horizontal='left')

def get_execution_status(sheet):
    """
    This function returns testexecution status for a test_type
    :param sheet: sheet object
    :param device_serial: device serial
    :param test_type: test_type
    :return: test execution data
    """
    runExcelRules(sheet)
    status = get_status_list(sheet)
    total_cases = len(status)
    passed_cases = status.count('Pass')
    failed_cases = status.count('Fail')
    not_executed_cases = total_cases - (passed_cases + failed_cases)
    print("*******get_exec_status_excel_report**********")
    execution_status_data = [total_cases, passed_cases, failed_cases, not_executed_cases]
    return execution_status_data